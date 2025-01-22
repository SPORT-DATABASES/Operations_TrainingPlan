import requests
from io import StringIO
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta, timezone
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Function to adjust timestamps and convert to local time
def convert_to_time(timestamp_ms, offset_hours=11):
    try:
        if pd.notnull(timestamp_ms):
            timestamp_s = float(timestamp_ms) / 1000
            return (datetime.fromtimestamp(timestamp_s, tz=timezone.utc) - timedelta(hours=offset_hours)).strftime('%H:%M')
    except (ValueError, TypeError):
        return None

# Function to format session strings with "and" and tab times under venues
def format_session_with_tabbed_time(session):
    if not session:
        return ""
    session = session.replace(' + ', ' and\n')  # Replace '+' with 'and\n' for better readability
    lines = session.split('\n')
    if len(lines) > 1:
        lines[-1] = f"\t{lines[-1]}"  # Add tab before the time
    return '\n'.join(lines)

# Function to format session information for a group
def format_session(group):
    venues = []
    times = set()
    for _, row in group.iterrows():
        venues.append(str(row['Venue']) if pd.notnull(row['Venue']) else '')  # Ensure Venue is a string
        times.add(f"{str(row['Start_Time']) if pd.notnull(row['Start_Time']) else ''}-"
                  f"{str(row['Finish_Time']) if pd.notnull(row['Finish_Time']) else ''}")
    venue_str = ' + '.join(filter(None, venues))  # Join only non-empty venues
    time_str = "\n".join(sorted(filter(None, times))) if times else ''
    return f"{venue_str}\n{time_str}"


# Function to ensure all expected columns are present in the pivot DataFrame
def ensure_all_columns(pivot_df, day_order):
    return pivot_df.reindex(columns=['Sport', 'Training_Group'] + day_order, fill_value=' ')

# Function to paste filtered data into the Template sheet
def paste_filtered_data_to_template(pivot_df, workbook, sport, training_group, start_cell):
    # Validate start_cell format (e.g., "C12")
    if not start_cell[0].isalpha() or not start_cell[1:].isdigit():
        raise ValueError(f"Invalid start_cell format: '{start_cell}'. Must be like 'C12'.")

    # Filter the DataFrame for the specified Sport and Training Group
    filtered_row = pivot_df[
        (pivot_df['Sport'] == sport) & (pivot_df['Training_Group'] == training_group)
    ]

    # Check if the filtered row is not empty
    if not filtered_row.empty:
        # Extract the values to paste (excluding Sport and Training_Group columns)
        values_to_paste = filtered_row.iloc[0, 2:].tolist()  # Skip first two columns

        # Get the starting row and column from the start_cell
        template_sheet = workbook["Template"]
        col_letter, row_num = start_cell[0], int(start_cell[1:])
        start_col_idx = ord(col_letter.upper()) - ord("A") + 1

        # Paste values into the Template sheet
        for col_idx, value in enumerate(values_to_paste, start=start_col_idx):
            cell = template_sheet.cell(row=row_num, column=col_idx, value=value)
            # Set alignment to center both vertically and horizontally
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        print(f"No data found for Sport='{sport}' and Training_Group='{training_group}'.")

def paste_concatenated_data(pivot_df, workbook, sport, start_cell):
    """
    Concatenate all text in each column for the specified sport in pivot_df,
    and paste the results into the Template sheet starting from the given cell.

    Args:
        pivot_df (DataFrame): The pivot DataFrame with the data.
        workbook (Workbook): The openpyxl workbook object.
        sport (str): The value to filter in the 'Sport' column.
        start_cell (str): The starting cell in the Template sheet (e.g., 'C47').
    """
    # Filter the DataFrame for the specified Sport
    filtered_df = pivot_df[pivot_df['Sport'] == sport]

    if filtered_df.empty:
        print(f"No data found for Sport='{sport}'.")
        return

    # Concatenate all text in each column (excluding 'Sport' and 'Training_Group')
    concatenated_values = filtered_df.iloc[:, 2:].apply(lambda col: "\n".join(col.dropna()), axis=0).tolist()

    # Get the starting row and column from the start_cell
    template_sheet = workbook["Template"]
    col_letter, row_num = start_cell[0], int(start_cell[1:])
    start_col_idx = ord(col_letter.upper()) - ord("A") + 1

    # Paste concatenated values into the Template sheet
    for col_idx, value in enumerate(concatenated_values, start=start_col_idx):
        cell = template_sheet.cell(row=row_num, column=col_idx, value=value)
        # Set alignment to center both vertically and horizontally
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

####################################################################################

# Fetch and parse the report
session = requests.Session()
session.auth = ("kenneth.mcmillan", "Quango76")
response = session.get("https://aspire.smartabase.com/aspireacademy/live?report=PYTHON3_TRAINING_PLAN&updategroup=true")
response.raise_for_status()
soup = BeautifulSoup(response.text, 'html.parser')
data = pd.read_html(StringIO(response.text))[0]
df = data.drop(columns=['About'], errors='ignore').drop_duplicates()

df.columns = df.columns.str.replace(' ', '_')  # Replace spaces in column headers
df['Group'] = df.apply(
    lambda row: f"{row['Sport']}-{row['Coach']}" 
    if row['Sport'] == row['Training_Group'] 
    else f"{row['Sport']}-{row['Training_Group']}-{row['Coach']}", 
    axis=1
)

df['Start_Time'] = pd.to_numeric(df['Start_Time'], errors='coerce').apply(lambda x: convert_to_time(x))
df['Finish_Time'] = pd.to_numeric(df['Finish_Time'], errors='coerce').apply(lambda x: convert_to_time(x))
# Drop rows where 'Sport' is blank (NaN or empty string)
df = df[df['Sport'].notna() & (df['Sport'].str.strip() != '')]

df = df[df['Venue'] != 'AASMC']

today = datetime.now()
next_sunday = today + timedelta(days=(6 - today.weekday()) % 7)
next_saturday = next_sunday + timedelta(days=6)
df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
df = df[(df['Date'] >= next_sunday.date()) & (df['Date'] <= next_saturday.date())]
unique_dates = sorted(df['Date'].dropna().unique())
df = (df.dropna(subset=['Sport'])
        .query("Sport.str.strip() != ''", engine='python')
        .assign(
            Date_long=lambda x: x['Date'].apply(lambda d: d.strftime('%a %d %b %Y') if pd.notnull(d) else None),
          #  Start_Time=lambda x: pd.to_numeric(x['Start_Time'], errors='coerce').apply(lambda t: convert_to_time(t)),
           # Finish_Time=lambda x: pd.to_numeric(x['Finish_Time'], errors='coerce').apply(lambda t: convert_to_time(t))
        )
        .drop(columns=['Date_Reverse'], errors='ignore')
        .sort_values(by=['Date', 'Sport', 'Coach', 'AM/PM'])
        .reset_index(drop=True)
)
df['AM/PM'] = pd.Categorical(df['AM/PM'], categories=['AM', 'PM'], ordered=True)
#df = df[df['by'] != 'Fusion Support']
df = df[['Date_long'] + [col for col in df.columns if col != 'Date_long']]
df = df[~df['Day_AM/PM'].str.contains('Friday', na=False)]

# Group and pivot data
grouped = df.groupby(['Sport', 'Training_Group', 'Day_AM/PM']).apply(
    lambda group: format_session(group)
).reset_index()

grouped.columns = ['Sport', 'Training_Group', 'Day_AM/PM', 'Session']
pivot_df = pd.pivot_table(
    grouped,
    values='Session',
    index=['Sport', 'Training_Group'],
    columns=['Day_AM/PM'],
    aggfunc='first',
    fill_value=' '
).reset_index()

# Ensure all day/time columns are present
day_order = [
    f"{day} {time}" for day in ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    for time in ['AM', 'PM']
]
pivot_df = ensure_all_columns(pivot_df, day_order)
pivot_df = pivot_df.applymap(lambda x: format_session_with_tabbed_time(x) if isinstance(x, str) else x)

# Load the template and paste data
template_path = "Excel_template.xlsx"
output_filename = f"{next_sunday.strftime('%d%b')}_automated_email_update.xlsx"
output_path = output_filename
if os.path.exists(output_path):
    os.remove(output_path)
shutil.copy(template_path, output_path)
workbook = load_workbook(output_path)

template_sheet = workbook["Template"]
date_cells = ['C4', 'E4', 'G4', 'I4', 'K4', 'M4', 'O4']

# Example rows to paste
    # Insert data into the template
rows_to_paste = [
    {"sport": "Development", "training_group": "Development 1", "start_cell": "C6"},
    {"sport": "Development", "training_group": "Development 2", "start_cell": "C8"},
    {"sport": "Development", "training_group": "Development 3", "start_cell": "C10"},
    {"sport": "Endurance", "training_group": "Endurance_Senior", "start_cell": "C12"},
    {"sport": "Jumps", "training_group": "Jumps_PV", "start_cell": "C14"},
    {"sport": "Jumps", "training_group": "Jumps_Martin Bercel", "start_cell": "C16"},
    {"sport": "Jumps", "training_group": "Jumps_Ross Jeffs", "start_cell": "C18"},
    {"sport": "Jumps", "training_group": "Jumps_ElWalid", "start_cell": "C20"},
    {"sport": "Sprints", "training_group": "Sprints_Lee", "start_cell": "C22"},
    {"sport": "Sprints", "training_group": "Sprints_Hamdi", "start_cell": "C24"},
    {"sport": "Throws", "training_group": "Performance Throws", "start_cell": "C26"},
    
    {"sport": "Squash", "training_group": "Squash", "start_cell": "C37"},
    {"sport": "Table Tennis", "training_group": "Table Tennis", "start_cell": "C39"},
    {"sport": "Fencing", "training_group": "Fencing", "start_cell": "C41"},
    {"sport": "Swimming", "training_group": "Swimming", "start_cell": "C43"},
    {"sport": "Padel", "training_group": "Padel", "start_cell": "C45"},
    
    #preacedemy padel is using concatenated function below C47
    
    {"sport": "Pre Academy", "training_group": "Pre Academy Fencing", "start_cell": "C49"},
    {"sport": "Pre Academy", "training_group": "Pre Academy Squash Girls", "start_cell": "C51"},
    {"sport": "Pre Academy", "training_group": "Pre Academy Athletics", "start_cell": "C53"},
    
        #athletics girls  is using concatenated function below C55
        
    {"sport": "Sprints", "training_group": "Sprints_Short", "start_cell": "C64"},
    {"sport": "Sprints", "training_group": "Sprints_Long", "start_cell": "C66"},
    
]

for row in rows_to_paste:
    paste_filtered_data_to_template(
        pivot_df=pivot_df,
        workbook=workbook,
        sport=row["sport"],
        training_group=row["training_group"],
        start_cell=row["start_cell"],
    )

paste_concatenated_data(pivot_df, workbook, sport="Pre Academy Padel", start_cell="C47")
paste_concatenated_data(pivot_df, workbook, sport="Girls Programe", start_cell="C55")

# Add dates to the template
date_cells = ['C4', 'E4', 'G4', 'I4', 'K4', 'M4', 'O4']
for idx, cell in enumerate(date_cells):
    day_offset = idx
    template_sheet[cell].value = (next_sunday + timedelta(days=day_offset)).strftime('%a %d %b %Y')
    template_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")

# Save the workbook
workbook.save(output_path)
print(f"Data successfully saved to the workbook: {output_path}")

###################################################################################

sender_email = "kennymcmillan29@gmail.com"
receiver_emails = ["kenneth.mcmillan@aspire.qa", "kennymcmillan29@gmail.com"]
subject = "Weekly Athletics Plan"
body = "Hi Alessandra,\n\nPlease find attached this week's Excel training plan. This is an automated email !\n\nBest Regards,\nAspire Data Team"

password = "lcsc pcuy pgxb zcri"

output_filename = output_path  # example

msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = ", ".join(receiver_emails)
msg['Subject'] = subject

msg.attach(MIMEText(body, 'plain'))

# Attach the Excel file
with open(output_filename, "rb") as attachment:
    part = MIMEApplication(attachment.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.add_header('Content-Disposition', 'attachment', filename=output_filename)
    msg.attach(part)

# Send the email via Gmail SMTP
with smtplib.SMTP('smtp.gmail.com', 587) as server:  # Using Gmail's SMTP server
    server.starttls()
    server.login(sender_email, password)
    server.send_message(msg)

print('Excel report generated and emailed successfully.')

