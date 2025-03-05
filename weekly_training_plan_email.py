import requests
from io import StringIO
import pandas as pd
from datetime import datetime, timedelta, timezone
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

###############################################################################
# Function to adjust timestamps and convert to local time
def convert_to_time(timestamp_ms, offset_hours=11):
    try:
        if pd.notnull(timestamp_ms):
            timestamp_s = float(timestamp_ms) / 1000
            return (datetime.fromtimestamp(timestamp_s, tz=timezone.utc) - timedelta(hours=offset_hours)).strftime('%H:%M')
    except (ValueError, TypeError):
        return None

###############################################################################
# Function to format session strings with "and" and tab times under venues
def format_session_with_tabbed_time(session):
    if not session:
        return ""
    session = session.replace(' + ', ' and\n')  # Replace '+' with 'and\n' for readability
    lines = session.split('\n')
    if len(lines) > 1:
        lines[-1] = f"\t{lines[-1]}"  # Add tab before the time
    return '\n'.join(lines)

###############################################################################
# Function to format session information for a group
###############################################################################
# Function to format session information for a group
def format_session(group):
    venue_time_pairs = []
    for _, row in group.iterrows():
        type_value = str(row['Session_Type']) if pd.notnull(row['Session_Type']) else ''  # Ensure Type is a string
        venue = str(row['Venue']) if pd.notnull(row['Venue']) else ''  # Ensure Venue is a string
        start_time = str(row['Start_Time']) if pd.notnull(row['Start_Time']) else ''
        finish_time = str(row['Finish_Time']) if pd.notnull(row['Finish_Time']) else ''
        time = f"{start_time}-{finish_time}" if start_time or finish_time else ''

        # If the session type is "Training Camp", return only "TRAINING CAMP"
        if type_value == "Training Camp":
            return "TRAINING CAMP"

        # Include "Competition" if Type is "Competition"
        if type_value == "Competition":
            formatted_entry = f"Competition\n{venue}\n{time}".strip()
        else:
            formatted_entry = f"{venue}\n{time}".strip()

        if venue or time or type_value:  # Include only non-empty entries
            venue_time_pairs.append((start_time, formatted_entry))

    # Sort the venue-time pairs by the start time
    sorted_venue_time_pairs = sorted(
        venue_time_pairs,
        key=lambda x: datetime.strptime(x[0], '%H:%M') if x[0] else datetime.min
    )

    # Extract only the formatted strings
    sorted_sessions = [pair[1] for pair in sorted_venue_time_pairs]

    # Join each entry with a single newline
    return '\n'.join(filter(None, sorted_sessions))



###############################################################################
# Function to ensure all expected columns are present in the pivot DataFrame
def ensure_all_columns(pivot_df, day_order):
    return pivot_df.reindex(columns=['Sport', 'Training_Group'] + day_order, fill_value=' ')

###############################################################################
# Function to paste filtered data into the Template sheet
# Added `no_data_found_list` to collect missing-data messages
def paste_filtered_data_to_template(pivot_df, workbook, sport, training_group, start_cell, no_data_found_list=None):
    # Validate start_cell format (e.g., "C12")
    if not start_cell[0].isalpha() or not start_cell[1:].isdigit():
        raise ValueError(f"Invalid start_cell format: '{start_cell}'. Must be like 'C12'.")

    # Filter the DataFrame for the specified Sport and Training Group
    filtered_row = pivot_df[
        (pivot_df['Sport'] == sport) & (pivot_df['Training_Group'] == training_group)
    ]

    # Check if the filtered row is not empty
    if not filtered_row.empty:
        values_to_paste = filtered_row.iloc[0, 2:].tolist()  # Skip first two columns
        template_sheet = workbook["Template"]
        col_letter, row_num = start_cell[0], int(start_cell[1:])
        start_col_idx = ord(col_letter.upper()) - ord("A") + 1

        # Paste values
        for col_idx, value in enumerate(values_to_paste, start=start_col_idx):
            cell = template_sheet.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        msg = f"{sport} - {training_group}."
        print(msg)
        if no_data_found_list is not None:
            no_data_found_list.append(msg)

###############################################################################
# Function to paste concatenated data for a single Sport
# Added `no_data_found_list` to collect missing-data messages
def paste_concatenated_data(pivot_df, workbook, sport, start_cell, no_data_found_list=None):
    # Filter the DataFrame for the specified Sport
    filtered_df = pivot_df[pivot_df['Sport'] == sport]

    if filtered_df.empty:
        msg = f"No data found for Sport='{sport}'."
        print(msg)
        if no_data_found_list is not None:
            no_data_found_list.append(msg)
        return

    # Concatenate all text in each column (excluding 'Sport' and 'Training_Group')
    concatenated_values = filtered_df.iloc[:, 2:].apply(lambda col: "\n".join(col.dropna()), axis=0).tolist()

    template_sheet = workbook["Template"]
    col_letter, row_num = start_cell[0], int(start_cell[1:])
    start_col_idx = ord(col_letter.upper()) - ord("A") + 1

    for col_idx, value in enumerate(concatenated_values, start=start_col_idx):
        cell = template_sheet.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

###############################################################################
# Main Script

# Fetch and parse the report
session = requests.Session()
session.auth = ("sb_sap.etl", "A1s2p3!re")  # Adjust if needed
response = session.get("https://aspire.smartabase.com/aspireacademy/live?report=PYTHON6_TRAINING_PLAN&updategroup=true")
response.raise_for_status()

data = pd.read_html(StringIO(response.text))[0]
df = data.drop(columns=['About'], errors='ignore').drop_duplicates()

df.columns = df.columns.str.replace(' ', '_')  # Replace spaces in column headers
df['Start_Time'] = pd.to_numeric(df['Start_Time'], errors='coerce').apply(lambda x: convert_to_time(x))
df['Finish_Time'] = pd.to_numeric(df['Finish_Time'], errors='coerce').apply(lambda x: convert_to_time(x))

# Drop rows where 'Sport' is blank (NaN or empty string)
df = df[df['Sport'].notna() & (df['Sport'].str.strip() != '')]

# Exclude this venue
df = df[df['Venue'] != 'AASMC']
df = df[df['Sport'] != 'Generic Athlete']
df = df[df['Training_Group'] != 'Practice']

# Define date range for the next week
today = datetime.now()
next_sunday = today + timedelta(days=(6 - today.weekday()) % 7)
next_saturday = next_sunday + timedelta(days=6)

df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
df = df[(df['Date'] >= next_sunday.date()) & (df['Date'] <= next_saturday.date())]

# Group and pivot data
grouped = (
    df.groupby(['Sport', 'Training_Group', 'Day_AM/PM', 'Session_Type'])  # Include Type in the grouping
    .apply(format_session)
    .reset_index()
)
grouped.columns = ['Sport', 'Training_Group', 'Day_AM/PM', 'Session_Type', 'Session']  # Update column names

# Create pivot table
pivot_df = pd.pivot_table(
    grouped,
    values='Session',
    index=['Sport', 'Training_Group'],
    columns=['Day_AM/PM'],
    aggfunc='first',
    fill_value=' '
).reset_index()

# Ensure all day/time columns are present
day_order = [
    f"{day} {time}"
    for day in ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    for time in ['AM', 'PM']
]
pivot_df = ensure_all_columns(pivot_df, day_order)

# Apply the tabbed-time format
pivot_df = pivot_df.applymap(lambda x: format_session_with_tabbed_time(x) if isinstance(x, str) else x)

# 3) Load the template and create an output Excel
template_path = "Excel_template.xlsx"
output_filename = f"{next_sunday.strftime('%d%b')}_{next_saturday.strftime('%d%b')}.xlsx"
output_path = output_filename

if os.path.exists(output_path):
    os.remove(output_path)

shutil.copy(template_path, output_path)
workbook = load_workbook(output_path)
template_sheet = workbook["Template"]

# Fill in date cells
# Add dates to the template
date_cells_groups = [
    ['C4', 'E4', 'G4', 'I4', 'K4', 'M4', 'O4'],  # Row 4
    ['C35', 'E35', 'G35', 'I35', 'K35', 'M35', 'O35'],  # Row 35
    ['C67', 'E67', 'G67', 'I67', 'K67', 'M67', 'O67'],  # Row 67
]

# Iterate through the date cells in groups
for day_offset, cell_group in enumerate(zip(*date_cells_groups)):  # Transpose groups
    date_value = (next_sunday + timedelta(days=day_offset)).strftime('%a %d %b %Y')  # Calculate the date
    for cell in cell_group:  # Assign the same date to the group
        template_sheet[cell].value = date_value
        template_sheet[cell].alignment = Alignment(horizontal="center", vertical="center")


# Add "Week Beginning" text to specific cells
week_number = next_sunday.isocalendar()[1]  # Get the ISO week number
week_beginning_text = f"Week beginning {next_sunday.strftime('%d %b')}\nWeek {week_number}"


# Populate cells O2, O33, and O65 with the week information
template_sheet["O2"].value = week_beginning_text
template_sheet["O33"].value = week_beginning_text
template_sheet["O65"].value = week_beginning_text

# Center align the text in these cells
for cell in ["O2", "O33", "O65"]:
    template_sheet[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 4) Paste data - Collect missing data messages in a list
no_data_found_messages = []

rows_to_paste = [
    {"sport": "Development", "training_group": "Development 1", "start_cell": "C6", "athlete_count": 14},
    {"sport": "Development", "training_group": "Development 2", "start_cell": "C8", "athlete_count": 11},
    {"sport": "Development", "training_group": "Development 3", "start_cell": "C10", "athlete_count": 9},
    {"sport": "Endurance", "training_group": "Endurance_Senior", "start_cell": "C12", "athlete_count": 18},
    {"sport": "Jumps", "training_group": "Jumps_Jaco", "start_cell": "C14", "athlete_count": 5},
    {"sport": "Jumps", "training_group": "Jumps_Martin Bercel", "start_cell": "C16", "athlete_count": 5},
    {"sport": "Jumps", "training_group": "Jumps_Ross Jeffs", "start_cell": "C18", "athlete_count": 6},
    {"sport": "Jumps", "training_group": "Jumps_ElWalid", "start_cell": "C20", "athlete_count": 9},
    {"sport": "Sprints", "training_group": "Sprints_Lee", "start_cell": "C22", "athlete_count": 8},
    {"sport": "Sprints", "training_group": "Sprints_Hamdi", "start_cell": "C24", "athlete_count": 9},
    {"sport": "Throws", "training_group": "Senior Performance Throws", "start_cell": "C26", "athlete_count": 12},
    {"sport": "Squash", "training_group": "Squash", "start_cell": "C37", "athlete_count": 13},
    {"sport": "Table Tennis", "training_group": "Table Tennis", "start_cell": "C39", "athlete_count": 5},
    {"sport": "Fencing", "training_group": "Fencing", "start_cell": "C41", "athlete_count": 16},
    {"sport": "Swimming", "training_group": "Swimming", "start_cell": "C43", "athlete_count": 16},
    {"sport": "Padel", "training_group": "Padel", "start_cell": "C45", "athlete_count": 9},
    {"sport": "Pre Academy Padel", "training_group": "Explorers", "start_cell": "C48", "athlete_count": 10},
    {"sport": "Pre Academy Padel", "training_group": "Explorers+", "start_cell": "C49", "athlete_count": 10},
    {"sport": "Pre Academy Padel", "training_group": "Starters", "start_cell": "C50", "athlete_count": 10},
    {"sport": "Pre Academy", "training_group": "Pre Academy Fencing", "start_cell": "C51", "athlete_count": 10},
    {"sport": "Pre Academy", "training_group": "Pre Academy Squash Girls", "start_cell": "C53", "athlete_count": 10},
    {"sport": "Pre Academy", "training_group": "Pre Academy Athletics", "start_cell": "C55", "athlete_count": 10},
    {"sport": "Girls Programe", "training_group": "Kids", "start_cell": "C58", "athlete_count": 16},
    {"sport": "Girls Programe", "training_group": "Mini Cadet_U14", "start_cell": "C59", "athlete_count": 8},
    {"sport": "Girls Programe", "training_group": "Cadet_U16", "start_cell": "C60", "athlete_count": 6},
    {"sport": "Girls Programe", "training_group": "Youth_U18", "start_cell": "C61", "athlete_count": 3},

    {"sport": "Sprints", "training_group": "Sprints_Steve", "start_cell": "C69", "athlete_count": 11},
    {"sport": "Sprints", "training_group": "Sprints_Kurt", "start_cell": "C71", "athlete_count": 14},
    {"sport": "Sprints", "training_group": "Sprints_Rafal", "start_cell": "C73", "athlete_count": 10},
    {"sport": "Sprints", "training_group": "Sprints_Francis", "start_cell": "C75", "athlete_count": 3},
    {"sport": "Endurance", "training_group": "Sprints_Yasmani", "start_cell": "C77", "athlete_count": 8},

    {"sport": "Endurance", "training_group": "Endurance_Driss", "start_cell": "C81", "athlete_count": 10},
    {"sport": "Endurance", "training_group": "Endurance_Kada", "start_cell": "C83", "athlete_count": 5},
    {"sport": "Endurance", "training_group": "Endurance_Khamis", "start_cell": "C85", "athlete_count": 11},
    {"sport": "Decathlon", "training_group": "Decathlon_QAF", "start_cell": "C87", "athlete_count": 7},
    
    {"sport": "Jumps", "training_group": "Jumps_Linus", "start_cell": "C96", "athlete_count": 4},
    {"sport": "Jumps", "training_group": "Jumps_Pawel", "start_cell": "C98", "athlete_count": 4},

    {"sport": "Throws", "training_group": "Discus_QAF", "start_cell": "C102", "athlete_count": 8},
    {"sport": "Throws", "training_group": "Hammer_QAF", "start_cell": "C104", "athlete_count": 4},
    {"sport": "Throws", "training_group": "Javelin_QAF", "start_cell": "C106", "athlete_count": 3},
]

# Paste row-by-row data
for row in rows_to_paste:
    paste_filtered_data_to_template(
        pivot_df=pivot_df,
        workbook=workbook,
        sport=row["sport"],
        training_group=row["training_group"],
        start_cell=row["start_cell"],
        no_data_found_list=no_data_found_messages
    )

# # Paste concatenated data
# paste_concatenated_data(
#     pivot_df=pivot_df,
#     workbook=workbook,
#     sport="Pre Academy Padel",
#     start_cell="C47",
#     no_data_found_list=no_data_found_messages
# )

# paste_concatenated_data(
#     pivot_df=pivot_df,
#     workbook=workbook,
#     sport="Girls Programe",
#     start_cell="C55",
#     no_data_found_list=no_data_found_messages
# )

# 5) Save the Excel workbook
workbook.save(output_path)
print(f"Data successfully saved to the workbook: {output_path}")

# 6) Build the email body to include "no data found" messages
body = (
    "Hi Alessandra,\n\n"
    "Please find attached this week's Excel training plan. This is an automated email. \n\n"
    "Best Regards,\nKenny"
)

# If there are missing-data messages, append them to the body
if no_data_found_messages:
    body += "\n\nThe following sports/groups had no data:\n"
    for msg in no_data_found_messages:
        body += f"- {msg}\n"

print("\nDebug: No data found messages:")
for m in no_data_found_messages:
    print("  ", m)

# 7) Send the email with the attached Excel
sender_email = "kennymcmillan29@gmail.com"
receiver_emails = ["kenneth.mcmillan@aspire.qa",
                   "kennymcmillan29@gmail.com",
                    "alessandra.moretti@aspire.qa"
]

subject = "Weekly Training Plan - Automated Thursday Update"

password = "lcsc pcuy pgxb zcri"

msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = ", ".join(receiver_emails)
msg['Subject'] = subject

msg.attach(MIMEText(body, 'plain'))

# Attach the Excel file
with open(output_path, "rb") as attachment:
    part = MIMEApplication(
        attachment.read(),
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    part.add_header('Content-Disposition', 'attachment', filename=output_filename)
    msg.attach(part)

# Send the email via Gmail SMTP
with smtplib.SMTP('smtp.gmail.com', 587) as server:
    server.starttls()
    server.login(sender_email, password)
    server.send_message(msg)

print('Excel report generated and emailed successfully.')
